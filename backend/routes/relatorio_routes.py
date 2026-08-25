import csv
import io

from flask import Blueprint, Response, jsonify, request

from services.agendamento_service import AgendamentoService
from routes.auth_routes import token_obrigatorio

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from routes.auth_routes import token_obrigatorio


relatorio_bp = Blueprint("relatorios", __name__)
service = AgendamentoService()


def _periodo():
    """Le inicio e fim da query string (?inicio=AAAA-MM-DD&fim=AAAA-MM-DD)."""
    inicio = request.args.get("inicio") or None
    fim = request.args.get("fim") or None

    return inicio, fim


@relatorio_bp.route(
    "/api/relatorios/faturamento",
    methods=["GET"],
)
@token_obrigatorio
def faturamento():
    inicio, fim = _periodo()

    return jsonify(
        service.relatorio_faturamento(inicio, fim)
    )


@relatorio_bp.route(
    "/api/relatorios/faturamento/csv",
    methods=["GET"],
)
def faturamento_csv():
    inicio, fim = _periodo()

    dados = service.relatorio_faturamento(inicio, fim)

    saida = io.StringIO()

    # O Excel em portugues espera ponto-e-virgula como
    # separador de colunas.
    escritor = csv.writer(saida, delimiter=";")

    escritor.writerow([
        "Profissional",
        "Atendimentos",
        "Faturamento",
    ])

    for linha in dados:
        valor = f"{linha['faturamento']:.2f}".replace(".", ",")

        escritor.writerow([
            linha["profissional"],
            linha["atendimentos"],
            valor,
        ])

    conteudo = "\ufeff" + saida.getvalue()

    nome = "faturamento.csv"

    if inicio or fim:
        nome = (
            f"faturamento_{inicio or 'inicio'}"
            f"_a_{fim or 'hoje'}.csv"
        )

    return Response(
        conteudo,
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={nome}"
        },
    )


@relatorio_bp.route(
    "/api/relatorios/faturamento/excel",
    methods=["GET"],
)
@token_obrigatorio
def faturamento_excel():
    inicio, fim = _periodo()

    dados = service.relatorio_faturamento(inicio, fim)

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Faturamento"

    # Cores
    roxo = "7A1FA2"
    roxo_claro = "F3E9F8"
    cinza = "B7B7B7"
    branco = "FFFFFF"

    # Titulo
    periodo = f"{inicio or 'inicio'} a {fim or 'hoje'}"

    planilha.merge_cells("A1:C1")

    titulo = planilha["A1"]
    titulo.value = f"Relatório de Faturamento - {periodo}"
    titulo.font = Font(
        bold=True,
        size=14,
        color=roxo,
    )
    titulo.alignment = Alignment(
        horizontal="center"
    )

    # Cabecalho
    cabecalho = [
        "Profissional",
        "Atendimentos",
        "Faturamento",
    ]

    preenchimento_roxo = PatternFill(
        fill_type="solid",
        fgColor=roxo,
    )

    fonte_branca = Font(
        bold=True,
        color=branco,
    )

    for coluna, valor in enumerate(cabecalho, start=1):
        celula = planilha.cell(
            row=3,
            column=coluna,
            value=valor,
        )

        celula.fill = preenchimento_roxo
        celula.font = fonte_branca
        celula.alignment = Alignment(
            horizontal="center"
        )

    # Borda inferior das linhas
    borda_inferior = Border(
        bottom=Side(
            style="thin",
            color=cinza,
        )
    )

    preenchimento_zebra = PatternFill(
        fill_type="solid",
        fgColor=roxo_claro,
    )

    # Dados
    linha_inicial = 4

    for indice, item in enumerate(dados, start=linha_inicial):
        planilha.cell(
            row=indice,
            column=1,
            value=item["profissional"],
        )

        planilha.cell(
            row=indice,
            column=2,
            value=item["atendimentos"],
        )

        celula_faturamento = planilha.cell(
            row=indice,
            column=3,
            value=item["faturamento"],
        )

        celula_faturamento.number_format = '"R$" #,##0.00'

        # Zebra nas linhas pares
        if (indice - linha_inicial) % 2 == 1:
            for coluna in range(1, 4):
                planilha.cell(
                    row=indice,
                    column=coluna,
                ).fill = preenchimento_zebra

        # Borda inferior
        for coluna in range(1, 4):
            planilha.cell(
                row=indice,
                column=coluna,
            ).border = borda_inferior

    # Linha de total
    linha_total = linha_inicial + len(dados)

    planilha.cell(
        row=linha_total,
        column=1,
        value="Total",
    )

    planilha.cell(
        row=linha_total,
        column=2,
        value=sum(
            item["atendimentos"]
            for item in dados
        ),
    )

    celula_total = planilha.cell(
        row=linha_total,
        column=3,
        value=sum(
            item["faturamento"]
            for item in dados
        ),
    )

    celula_total.number_format = '"R$" #,##0.00'

    borda_superior_dupla = Border(
        top=Side(
            style="double",
            color=roxo,
        )
    )

    for coluna in range(1, 4):
        celula = planilha.cell(
            row=linha_total,
            column=coluna,
        )

        celula.font = Font(
            bold=True,
            color=roxo,
        )

        celula.border = borda_superior_dupla

    # Larguras das colunas
    planilha.column_dimensions["A"].width = 32
    planilha.column_dimensions["B"].width = 16
    planilha.column_dimensions["C"].width = 18

    # Congelar cabecalho
    planilha.freeze_panes = "A4"

    # Salvar em memoria
    arquivo = io.BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)

    # Nome do arquivo
    nome = "faturamento.xlsx"

    if inicio or fim:
        nome = (
            f"faturamento_{inicio or 'inicio'}"
            f"_a_{fim or 'hoje'}.xlsx"
        )

    return Response(
        arquivo.getvalue(),
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f"attachment; filename={nome}"
            )
        },
    )

@relatorio_bp.route("/api/relatorios/faturamento",
methods=["GET"])
@token_obrigatorio
def faturamento():
 return jsonify(service.relatorio_faturamento())